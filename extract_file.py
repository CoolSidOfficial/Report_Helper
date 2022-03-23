import  openpyxl
import banner
import  os.path  
import time
import csv
from colorama import Fore ,Style,init,Back
init()
alldata=[]
#-_______________________________________________________________________________________________________________________
def gettingPathExcel():
     print("Please input your  excel file path  :) ")
     ex_path=input("[>>]")
     if not ex_path !="":
         print("You didn't gave me any file, how will I open it then :(")
         print("Try again")    
         gettingPathExcel()
     if not os.path.isfile(ex_path):
         print("File does not exist :(")
         print("TryAgain :)")
         gettingPathExcel()
     else:
       
        print("Opening",ex_path)
        return ex_path
#--------------------------------------------------------------------------------------------------------------------------------------------
def excelExtractor(path,column=2):
     excelfile=openpyxl.load_workbook(path)    
     sheet_activate=excelfile.active
     for i in range(1,sheet_activate.max_row+1):
         cellobj=sheet_activate.cell(row=i,column=column)
         alldata.append([i,cellobj.value])   # it will  put all data into a list 
        

def extractingCsv():
    with open("reports.csv","w",newline="") as file:
       writer=csv.writer(file,delimiter="|") 
       writer.writerows(alldata)

def importingCsv(row):
    if  not os.path.isfile("reports.csv"):
       print("Old file does not exist Upload it again")
       extractFile()
    with open("reports.csv","r") as file:
        reader=list(csv.reader(file,delimiter="|"))  
        finaldata=reader[row-1]  # to get the accurate number as list it from zero
        return finaldata[1]   # so we can only get description
#_________________________________________________________________________________________________________________________
def extractFile():
    print(banner.echo_banner)
    print(Fore.YELLOW+Style.BRIGHT+"Do you want to open the last  excel file  or Upload  a new Excel file".center(120)) 
    print("[1>>]Use Current Excel File \n[2>>]Upload a new Excel file\n")
    c1=input("Enter the choice number:#>>")
    if c1 == "" :
        print("You didn't enter anything ")
        print("Try Again")
        extractFile()
    if c1=="1" or c1=="one":
           return importingCsv(grabId())  # it will return that description
    if c1=="2" or c1=="two":
       path=gettingPathExcel()
       excelExtractor(path)
       extractingCsv()
       return importingCsv(grabId())         

def grabId():
     print("Please enter the report number you wish to copy :)".center(80))
     c3=input("[#>>]")
     if c3=="":
            print("Row number cannot be empty")
            time.sleep(1)
            extractFile()
     return int(c3)
     
if __name__=='__main__':
   print(extractFile())

# done 